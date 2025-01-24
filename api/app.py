from fastapi import FastAPI, File, UploadFile, HTTPException
from pydantic import BaseModel
from langchain_community.llms import Ollama
from langchain_community.embeddings import OllamaEmbeddings
from langchain.chains import RetrievalQA
from langchain.text_splitter import CharacterTextSplitter
from langchain_community.vectorstores import FAISS
import os
import time
import requests
from requests.exceptions import RequestException

app = FastAPI()

# Update Ollama connection URLs to use the service name
llm = Ollama(base_url="http://localhost:11434", model="llama2")
embeddings = OllamaEmbeddings(base_url="http://localhost:11434", model="llama2")

class QuestionRequest(BaseModel):
    question: str

@app.post("/upload")
async def upload_files(files: list[UploadFile] = File(...)):
    file_paths = []
    for file in files:
        file_path = f"./uploads/{file.filename}"
        os.makedirs(os.path.dirname(file_path), exist_ok=True)
        with open(file_path, "wb") as buffer:
            buffer.write(file.file.read())
        file_paths.append(file_path)
    
    # Tạo QA chain và lưu vector store
    qa_chain, vectorstore = create_qa_chain(file_paths)
    save_vectorstore(vectorstore)
    return {"message": "Files uploaded and processed successfully", "file_paths": file_paths}

@app.post("/ask")
async def ask_question(request: QuestionRequest):
    if not os.path.exists("vectorstore"):
        raise HTTPException(status_code=400, detail="No files have been uploaded yet.")
    
    # Tải lại vector store
    vectorstore = load_vectorstore()
    qa_chain = RetrievalQA.from_chain_type(
        llm=llm,
        chain_type="stuff",
        retriever=vectorstore.as_retriever()
    )
    
    # Trả lời câu hỏi
    response = qa_chain.run(request.question)
    return {"answer": response}

def create_qa_chain(file_paths):
    # Đọc nội dung từ nhiều file
    combined_text = read_files(file_paths)
    
    # Chia nhỏ văn bản thành các đoạn
    text_splitter = CharacterTextSplitter(chunk_size=1000, chunk_overlap=200)
    texts = text_splitter.split_text(combined_text)
    
    # Tạo embeddings và lưu vào vector store
    vectorstore = FAISS.from_texts(texts, embeddings)
    
    # Tạo QA chain
    qa_chain = RetrievalQA.from_chain_type(
        llm=llm,
        chain_type="stuff",
        retriever=vectorstore.as_retriever()
    )
    return qa_chain, vectorstore

def save_vectorstore(vectorstore, path="vectorstore"):
    vectorstore.save_local(path)

def load_vectorstore(path="vectorstore"):
    vectorstore = FAISS.load_local(
        path, 
        embeddings,
        allow_dangerous_deserialization=True
    )
    return vectorstore

def read_files(file_paths):
    combined_text = ""
    for file_path in file_paths:
        if file_path.endswith(".pdf"):
            from PyPDF2 import PdfReader
            reader = PdfReader(file_path)
            for page in reader.pages:
                combined_text += page.extract_text()
        elif file_path.endswith(".docx"):
            from docx import Document
            doc = Document(file_path)
            combined_text += "\n".join([para.text for para in doc.paragraphs])
        elif file_path.endswith(".xlsx"):
            from openpyxl import load_workbook
            wb = load_workbook(file_path)
            for sheet in wb:
                for row in sheet.iter_rows(values_only=True):
                    combined_text += " ".join([str(cell) for cell in row if cell is not None]) + "\n"
    return combined_text